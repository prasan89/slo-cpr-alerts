from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


HEADERS = [
    "timestamp_ist", "symbol", "ltp", "r3", "r2", "r1", "tc", "pivot", "bc",
    "s1", "s2", "s3", "yesterday_r1", "yesterday_s1", "r1_improving", "s1_improving",
    "cpr_width_pct", "state", "alert", "previous_ltp", "reference_session", "prior_session",
]
SIGNAL_HEADERS = ["timestamp_ist", "symbol", "signal", "ltp", "trigger_level"]


def append_snapshot(path: str | Path, row: dict) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        wb = load_workbook(path)
        ws = wb["CPR Alerts"]
        existing = [cell.value for cell in ws[1]]
        if existing != HEADERS:
            ws.delete_rows(1, ws.max_row)
            ws.append(HEADERS)
            for cell in ws[1]:
                cell.font = Font(bold=True)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "CPR Alerts"
        ws.append(HEADERS)
        for cell in ws[1]:
            cell.font = Font(bold=True)

    ws = wb["CPR Alerts"]
    ws.append([row.get(h) for h in HEADERS])
    ws.freeze_panes = "A2"
    wb.save(path)


def append_signal(path: str | Path, timestamp: str, symbol: str, signal: str, ltp: float, trigger_level: float) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(path) if path.exists() else Workbook()
    if "CPR Alerts" not in wb.sheetnames:
        ws = wb.active
        ws.title = "CPR Alerts"
    if "Signals" not in wb.sheetnames:
        ws = wb.create_sheet("Signals")
        ws.append(SIGNAL_HEADERS)
        for cell in ws[1]:
            cell.font = Font(bold=True)
    ws = wb["Signals"]
    ws.append([timestamp, symbol, signal, ltp, trigger_level])
    ws.freeze_panes = "A2"
    wb.save(path)


def create_workbook(path: str | Path) -> None:
    path = Path(path)
    if path.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "CPR Alerts"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws = wb.create_sheet("Signals")
    ws.append(SIGNAL_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    wb.create_sheet("Latest Snapshot")
    wb.create_sheet("CPR Levels")
    wb.save(path)
